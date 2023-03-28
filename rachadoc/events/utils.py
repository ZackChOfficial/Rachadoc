from events.models import Appointement
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Protection
from events.choices import Status
import datetime


def _get_excel_filename(from_date: datetime, to_date: datetime):
    filename = "export_rendez_vous"
    if from_date and to_date:
        filename += f"_entre_{from_date}_et_{to_date}"
    elif from_date:
        filename += f"_apres_{from_date}"
    elif to_date:
        filename += f"_avant_{to_date}"
    return filename


def _get_qs_appointement(from_date: datetime = None, to_date: datetime = None, **filters):
    qs = Appointement.objects.order_by("start").all()
    if from_date:
        qs = qs.filter(start__gte=from_date)
    if to_date:
        qs = qs.filter(end__lte=to_date)
    qs = qs.filter(**filters)
    return qs


def export_appointement_by_xls(from_date: datetime = None, to_date: datetime = None, **filters):
    try:
        qs = _get_qs_appointement(from_date, to_date, **filters)
        header = [
            "Prenom",
            "Nom",
            "Email",
            "Telephone",
            "Date de début",
            "Date de fin",
            "Status",
            "Note pré rendez-vous",
            "Note après le rendez-vous",
            "Creer par",
        ]
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title="liste des rendez vous", index=1)
        filename = _get_excel_filename(from_date, to_date)
        row_num = 1
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True)
        for __, appointement in enumerate(qs, 1):
            row_num += 1

            row = [
                appointement.patient.first_name,
                appointement.patient.last_name,
                appointement.patient.email,
                appointement.patient.phone_number,
                str(appointement.start),
                str(appointement.end),
                appointement.get_status_display(),
                appointement.note_pre_appointement,
                appointement.note_post_appointement,
                f"{appointement.creator.first_name} {appointement.creator.last_name} {appointement.creator.email}",
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.protection = Protection(locked=True)
        workbook.save(f"{filename}.xlsx")
    except Exception as e:
        print(e)
